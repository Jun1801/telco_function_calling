"""Parse data/Function.xlsx (26 real Viettel KPI functions) into tool registry files.

Outputs:
  data/real_tools.json           — 26 tools (tools.json format) + seed_examples + split,
                                    with reference codes injected as enums in the schema.
  data/real_reference_codes.json — location/kpi/unit code tables for the generator.
  data/real_station_catalogue.json — synthetic station catalogue (station_code → location).
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "Function.xlsx"

# 8 tools held out for unseen generalization eval.
UNSEEN_TOOLS = {
    "radio_kpi",
    "thong_ke_cntt",
    "top_tram_max",
    "top_tram_min",
    "top_cell_max",
    "top_cell_min",
    "top_sub_attached_max",
    "top_sub_attached_min",
}

# Params that must be constrained to a reference catalogue (enum injected into schema).
# object_code is intentionally NOT enum'd: per the Excel spec it accepts location OR
# station OR sector/device codes (open set).
_ENUM_PARAMS = {"location_code", "kpi_code", "unit_code", "station_code"}


def _parse_query_examples(cell: str | None) -> list[str]:
    if not cell:
        return []
    queries = []
    for line in str(cell).split("\n"):
        line = line.strip()
        if line.startswith("-"):
            line = line.lstrip("-").strip()
        if line:
            queries.append(line)
    return queries


def _parse_call_example(cell: str | None) -> dict | None:
    if not cell:
        return None
    try:
        return json.loads(cell)
    except (json.JSONDecodeError, TypeError):
        return None


def parse_reference_codes(wb) -> dict:
    loc_ws = wb["location_codeobject_code"]
    locations = []
    for group, value, code in list(loc_ws.iter_rows(values_only=True))[1:]:
        if value and code:
            locations.append({"code": code, "name": value, "group": group})

    kpi_ws = wb["kpi_code"]
    kpis = [
        {"code": c, "meaning": m}
        for c, m in list(kpi_ws.iter_rows(values_only=True))[1:]
        if c
    ]

    unit_ws = wb["unit_code"]
    units = [
        {"code": c, "name": n}
        for c, n in list(unit_ws.iter_rows(values_only=True))[1:]
        if c
    ]
    return {"location_code": locations, "kpi_code": kpis, "unit_code": units}


def build_station_catalogue(references: dict, per_location: int = 3, cap: int = 300) -> list[dict]:
    """Synthetic station catalogue for tools needing station_code (no real catalogue exists).

    Format EHB000NN matches the Excel seed (e.g. EHB00041). Each station maps to a
    province-level location so queries can name a station within a location.
    """
    # Provinces FIRST (keeps EHB00001.. mapping stable), then regions/countries/nationwide
    # so EVERY location the sampler uses has stations → mock observations stay
    # geographically consistent. Dedupe codes preserving order.
    seen, ordered = set(), []
    province_codes = {loc["code"] for loc in references["location_code"]
                      if loc.get("group") == "Tỉnh/Thành phố Việt Nam"}
    for loc in references["location_code"]:
        if loc["code"] in seen:
            continue
        if loc["code"] in province_codes:
            seen.add(loc["code"]); ordered.append(loc)
    for loc in references["location_code"]:  # regions, countries, nationwide
        if loc["code"] not in seen:
            seen.add(loc["code"]); ordered.append(loc)

    catalogue: list[dict] = []
    n = 1
    for loc in ordered:
        for k in range(per_location):
            if len(catalogue) >= cap:
                return catalogue
            catalogue.append({
                "station_code": f"EHB{n:05d}",
                "location_code": loc["code"],
                "name": f"Trạm {loc['name']} {k + 1}",
            })
            n += 1
    return catalogue


def _enum_values(param: str, references: dict, stations: list[dict]) -> tuple[list, str]:
    """Return (enum_values, short_hint) for a constrained param."""
    if param == "location_code":
        # Dedup codes preserving order.
        seen, vals, pairs = set(), [], []
        for loc in references["location_code"]:
            if loc["code"] not in seen:
                seen.add(loc["code"])
                vals.append(loc["code"])
                pairs.append(f"{loc['code']}={loc['name']}")
        hint = "Mã vị trí. Ví dụ: " + ", ".join(pairs[:6]) + ", …"
        return vals, hint
    if param == "kpi_code":
        vals = [k["code"] for k in references["kpi_code"]]
        pairs = [f"{k['code']}={k['meaning']}" for k in references["kpi_code"]]
        return vals, "Mã KPI. Gồm: " + "; ".join(pairs[:4]) + "; …"
    if param == "unit_code":
        vals = [u["code"] for u in references["unit_code"]]
        pairs = [f"{u['code']}={u['name']}" for u in references["unit_code"]]
        return vals, "Mã đơn vị. Gồm: " + "; ".join(pairs[:4]) + "; …"
    if param == "station_code":
        vals = [s["station_code"] for s in stations]
        return vals, "Mã trạm/cell. Ví dụ: " + ", ".join(vals[:5]) + ", …"
    return [], ""


def _inject_enums(parameters: dict, references: dict, stations: list[dict]) -> None:
    props = parameters.get("properties", {})
    for param, spec in props.items():
        if param in _ENUM_PARAMS and "enum" not in spec:
            vals, hint = _enum_values(param, references, stations)
            if vals:
                spec["enum"] = vals
                base = spec.get("description", "").strip()
                spec["description"] = (base + " — " + hint).strip(" —") if base else hint
        if param == "object_code" and "enum" not in spec:
            base = spec.get("description", "").strip()
            note = "Chấp nhận mã location (VNM/tỉnh/huyện) hoặc mã trạm."
            if note not in base:
                spec["description"] = (base + " " + note).strip()


def parse_functions(wb, references: dict, stations: list[dict]) -> list[dict]:
    ws = wb["Function"]
    rows = list(ws.iter_rows(values_only=True))
    tools = []
    for row in rows[1:]:
        if not row[0]:
            continue
        raw_name = str(row[0]).strip()
        name = raw_name.lower()
        schema_info = json.loads(row[3])
        parameters = schema_info["parameters"]
        parameters.setdefault("type", "object")
        parameters.setdefault("required", [])
        parameters.setdefault("properties", {})
        _inject_enums(parameters, references, stations)

        queries = _parse_query_examples(row[4])
        call_args = _parse_call_example(row[5])
        seed_examples = [
            {"query": q, "call": call_args} if call_args else {"query": q}
            for q in queries
        ]

        tools.append(
            {
                "name": name,
                "raw_name": raw_name,
                "domain": "kpi_reporting",
                "description": str(row[1]).strip() if row[1] else schema_info.get("description", ""),
                "deprecated": False,
                "status": "active",
                "risk_level": "low",
                "side_effect": "read",
                "side_effects": [],
                "parameters": parameters,
                "version": "1.0",
                "tool_id": f"telco.kpi.{name}.v1.0",
                "split": "unseen" if name in UNSEEN_TOOLS else "seen",
                "tool_group": "kpi_reporting",
                "permission_required": None,
                "dependencies": [],
                "seed_examples": seed_examples,
                "seed_call": call_args,
            }
        )
    return tools


def main() -> None:
    if not XLSX.exists():
        sys.exit(f"Missing {XLSX}")
    wb = openpyxl.load_workbook(XLSX)
    references = parse_reference_codes(wb)
    stations = build_station_catalogue(references)
    tools = parse_functions(wb, references, stations)
    # Real KPI tools are read-only → no contracts (no precondition/permission/side-effect).

    data_dir = ROOT / "data"
    (data_dir / "real_tools.json").write_text(json.dumps(tools, ensure_ascii=False, indent=2), encoding="utf-8")
    (data_dir / "real_reference_codes.json").write_text(json.dumps(references, ensure_ascii=False, indent=2), encoding="utf-8")
    (data_dir / "real_station_catalogue.json").write_text(json.dumps(stations, ensure_ascii=False, indent=2), encoding="utf-8")

    seen = sum(1 for t in tools if t["split"] == "seen")
    enum_tools = sum(1 for t in tools if any("enum" in p for p in t["parameters"]["properties"].values()))
    print(f"Parsed {len(tools)} tools ({seen} seen, {len(tools) - seen} unseen) → data/real_tools.json")
    print(f"References: {len(references['location_code'])} loc, {len(references['kpi_code'])} kpi, "
          f"{len(references['unit_code'])} unit | stations: {len(stations)} | tools with enum: {enum_tools}")


if __name__ == "__main__":
    main()
